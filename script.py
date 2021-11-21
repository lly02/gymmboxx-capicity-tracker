from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from datetime import datetime, time, timedelta
from openpyxl.utils import get_column_letter

def getData(driver):
    capacities, locations = [], []

    for k in range(10):
        try:
            for x in range(7):
                capacity = driver.find_element(By.XPATH, f"//div[@id='outlets']/div[{x+1}]/div/div[@class='bar-bg']/div").get_attribute("style")
                location = driver.find_element(By.XPATH, f"//div[@id='outlets']/div[{x+1}]/div/div[contains(@class, 'lead')]").text

                capacity = capacity.split(";")[0].split(":")[1]

                replaceChar = "% "

                for x in replaceChar:
                    capacity = capacity.replace(x, "")

                capacities.append(capacity)
                locations.append(location)

                break
        except Exception as e:
            if k == 9:
                driver.quit()
                raise Exception(e)

            pass

    driver.quit()
    
    return [capacities, locations]

def getTimeList(y, x):
    if(len(str(y)) < 2): 
        hour = f"0{y}"
    else:
        hour = y

    if(x == 1):
        min = "30"
    else:
        min = "00"

    return [hour, min]


def main():
    #selenium setup
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")

    driver = webdriver.Chrome(options=chrome_options)

    driver.get("https://smartentry.org/status/gymmboxx")

    #get data from website
    data = getData(driver)
    
    capacities = data[0] 
    locations = data[1]

    #load workbook/ create workbook is !exist
    try:
        wb = load_workbook("data.xlsx")
    except:
        wb = Workbook()
        
    #update if hv new locations/ create sheets for new excel
    wb.worksheets[0].title = "Summary"
    worksheetsCopy = wb.worksheets

    for x in locations:
        #check if location exists
        flag = True

        for y in range(len(worksheetsCopy)):
            if worksheetsCopy[y].title == x:
                flag = False
                del worksheetsCopy[y]

                break

        #create sheet and time row
        if flag:
            ws = wb.create_sheet(x)
            ws.column_dimensions["A"].width = 15
            
            cell = wb[x].cell(1,2)
            cell.value = "Average"

            cell.fill = PatternFill(fill_type="solid",
                                    start_color="00FFFF00")
            cell.border = Border(   left=Side(border_style = "thin"),
                                    top=Side(border_style = "thin"),
                                    right=Side(border_style = "thin"),
                                    bottom=Side(border_style = "thin"))

            wb[x].merge_cells("B1:AW1")

            for y in range(24):
                for k in range(2):
                    #avg
                    cell = wb[x].cell(2, 2)

                    if 66 + (k+y*2) <= 90:
                        cellChar = chr(66 + (k+y*2))
                    else:
                        cellChar = 'A' + chr(66 + (k+y*2) - 26)

                    cell = wb[x].cell(2, 2+k+y*2)
                    cell.value = f"=AVERAGE({cellChar}5:{cellChar}1048576)"
                    cell.border = Border(   left=Side(border_style = "thin"),
                                            top=Side(border_style = "thin"),
                                            right=Side(border_style = "thin"),
                                            bottom=Side(border_style = "thin"))

                    #time
                    cell = wb[x].cell(4, 2+k+y*2)

                    cell.number_format = "HH:MM"
                    cell.fill = PatternFill(fill_type="solid",
                                            start_color="00FFFF00")
                    cell.border = Border(   left=Side(border_style = "thin"),
                                            top=Side(border_style = "thin"),
                                            right=Side(border_style = "thin"),
                                            bottom=Side(border_style = "thin"))
                    
                    timeList = getTimeList(y, k)
                    
                    cell.value = f"{str(timeList[0])}:{timeList[1]}"

    #insert data
    timeNow = datetime.now()
    timeNow = timeNow - timedelta(minutes=timeNow.minute % 30)
    timeNow = timeNow.strftime("%H:%M")
    timeIndex = 0

    flag = False

    for x in range(24):
        for k in range(2):
            timeList = getTimeList(x, k)
            
            if timeNow == f"{str(timeList[0])}:{timeList[1]}":
                timeIndex = x*2+k+2
                flag = True
                break

        if flag: break

    locationCopy = locations

    for x in range(len(wb.sheetnames)):
        ws = wb[wb.sheetnames[x]]
        if ws.title == "Summary": continue

        for y in range(len(locationCopy)):
            flag = False

            if ws.title == locationCopy[y]:
                max_col = len(ws[get_column_letter(timeIndex)]) + 1

                if ws.cell(max_col, 1).value == None:
                    cell = ws.cell(max_col, 1)
                    cell.number_format = "DD MM YYYY"
                    cell.value = datetime.now().date().strftime("%d/%m/%Y")
                    cell.border = Border(   left=Side(border_style = "thin"),
                                            top=Side(border_style = "thin"),
                                            right=Side(border_style = "thin"),
                                            bottom=Side(border_style = "thin"))
                    cell.fill = PatternFill(fill_type="solid",
                                            start_color="00FFFF00")                              

                ws.cell(max_col, timeIndex).number_format = "0"
                ws.cell(max_col, timeIndex).value = int(capacities[y])
                
                del locationCopy[y]
                del capacities[y]

                flag = True
            
            if flag: break

    wb.save("data.xlsx")